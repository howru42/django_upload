from django.http import HttpResponse
from .models import Record
from .forms import UploadForm
import openpyxl
from django.views.decorators.csrf import csrf_exempt, csrf_protect


@csrf_exempt
def upload(request):
    form = UploadForm(request.POST, request.FILES)
    _str = "<H2>Failure</h2>"
    if (form.is_valid()):
        parseFile(request.FILES["file"])
        wb = openpyxl.load_workbook('example_out.xlsx')
        sheet = wb.get_sheet_by_name('Sheet1')
        Record.objects.all().delete()
        for i in range(1, sheet.max_row):
            if sheet.cell(row=i, column=3).value is not None:
                record = Record(name=str(sheet.cell(row=i, column=2).value), amount=sheet.cell(row=i, column=3).value)
                record.save()
                _str = _str + "<h2>" + record.name + " - " + str(record.amount) + "</h2><br/>"
    return HttpResponse(_str)


def parseFile(f):
    with open('example_out.xlsx', 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)
    return
